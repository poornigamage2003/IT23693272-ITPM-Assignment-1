from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
import argparse
import re
import sys
import time

DEFAULT_URL = "https://www.pixelssuite.com/chat-translator"
DEFAULT_SHEET_NAMES = ["Test cases", " Test cases"]

INPUT_HEADERS = ["Input", "Singlish", "Singlish Input", "Test Input", "Source", "Sentence", "Text"]
EXPECTED_HEADERS = ["Expected output", "Expected Output", "Expected_Output", "Expected", "Sinhala", "Expected Sinhala"]
ACTUAL_HEADERS = ["Actual output", "Actual Output", "Actual_Output", "Actual"]
STATUS_HEADERS = ["Status", "Result", "Pass/Fail", "Pass Fail"]


def configure_stdout():
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")
    except Exception:
        pass


def normalize_header(value):
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())


def find_header_row(ws, max_scan_rows=30):
    input_tokens = {normalize_header(v) for v in INPUT_HEADERS}
    expected_tokens = {normalize_header(v) for v in EXPECTED_HEADERS}

    for row in range(1, min(ws.max_row, max_scan_rows) + 1):
        values = [ws.cell(row=row, column=col).value for col in range(1, ws.max_column + 1)]
        norms = {normalize_header(v) for v in values if v is not None}
        if (norms & input_tokens) and (norms & expected_tokens):
            return row

    raise RuntimeError("Could not find the header row. Check that the Excel file has Input and Expected output columns.")


def find_column(headers, candidates):
    normalized = {
        normalize_header(value): index
        for index, value in enumerate(headers, start=1)
        if value is not None
    }

    for candidate in candidates:
        found = normalized.get(normalize_header(candidate))
        if found:
            return found

    return None


def ensure_column(ws, header_row, headers, preferred_name, candidates):
    existing = find_column(headers, candidates)
    if existing:
        return existing

    new_col = len(headers) + 1
    ws.cell(row=header_row, column=new_col).value = preferred_name
    headers.append(preferred_name)
    return new_col


def top_left_cell(ws, row, col):
    cell = ws.cell(row=row, column=col)

    if not isinstance(cell, MergedCell):
        return cell

    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return ws.cell(row=rng.min_row, column=rng.min_col)

    return cell


def is_top_left_or_normal(ws, row, col):
    cell = ws.cell(row=row, column=col)

    if not isinstance(cell, MergedCell):
        return True

    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return row == rng.min_row and col == rng.min_col

    return True


def pick_sheet(wb, requested_sheet=None):
    if requested_sheet and requested_sheet in wb.sheetnames:
        return wb[requested_sheet]

    for name in DEFAULT_SHEET_NAMES:
        if name in wb.sheetnames:
            return wb[name]

    return wb.active


def get_chat_locators(page):
    page.wait_for_selector("textarea", timeout=60000)

    input_box = page.locator('textarea[placeholder*="English"]').first
    output_box = page.locator('textarea[placeholder*="Sinhala"]').first

    # Fallback for layout changes: first textarea = input, second textarea = output.
    if input_box.count() == 0 or output_box.count() == 0:
        textareas = page.locator("textarea")
        if textareas.count() < 2:
            raise RuntimeError("Could not find both input and output textareas on the website.")
        input_box = textareas.nth(0)
        output_box = textareas.nth(1)

    button = page.get_by_role("button", name=re.compile(r"transliterate", re.IGNORECASE)).first
    if button.count() == 0:
        button = page.locator("button").last

    return input_box, output_box, button


def read_output_value(output_box):
    try:
        value = output_box.input_value(timeout=2000)
        if value:
            return value.strip()
    except Exception:
        pass

    try:
        value = output_box.inner_text(timeout=2000)
        if value:
            return value.strip()
    except Exception:
        pass

    return ""


def run_single_case(context, url, singlish_input, wait_ms, type_delay_ms, timeout_ms):
    """Open a fresh page for every test case to avoid stale/previous actual outputs."""
    page = context.new_page()
    page.set_default_timeout(timeout_ms)

    try:
        page.goto(url, wait_until="domcontentloaded", timeout=timeout_ms)
        page.wait_for_timeout(2500)

        input_box, output_box, button = get_chat_locators(page)

        input_box.click(timeout=5000)
        page.keyboard.press("Control+A")
        page.keyboard.press("Backspace")

        if type_delay_ms and type_delay_ms > 0:
            input_box.type(singlish_input, delay=type_delay_ms)
        else:
            input_box.fill(singlish_input)

        page.wait_for_timeout(600)

        try:
            button.click(force=True, timeout=5000)
        except Exception:
            page.keyboard.press("Enter")

        start_time = time.time()
        actual_output = ""

        while (time.time() - start_time) * 1000 < wait_ms:
            actual_output = read_output_value(output_box)
            if actual_output:
                return actual_output
            page.wait_for_timeout(500)

        return actual_output

    finally:
        page.close()


def run_tests(args):
    excel_path = Path(args.excel).resolve()

    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    print(f"Loading Workbook: {excel_path.name}")
    wb = load_workbook(excel_path)
    ws = pick_sheet(wb, args.sheet)

    header_row = find_header_row(ws)
    headers = [ws.cell(row=header_row, column=col).value for col in range(1, ws.max_column + 1)]

    input_col = find_column(headers, INPUT_HEADERS)
    expected_col = find_column(headers, EXPECTED_HEADERS)
    actual_col = ensure_column(ws, header_row, headers, "Actual output", ACTUAL_HEADERS)
    status_col = ensure_column(ws, header_row, headers, "Status", STATUS_HEADERS)

    if not input_col:
        raise RuntimeError("Input column not found.")
    if not expected_col:
        raise RuntimeError("Expected output column not found.")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=args.headless, slow_mo=args.slow_mo_ms)
        context = browser.new_context()

        tested_count = 0
        print(f"Opening: {args.url}")

        for row in range(header_row + 1, ws.max_row + 1):
            if not is_top_left_or_normal(ws, row, input_col):
                continue

            singlish_input = str(top_left_cell(ws, row, input_col).value or "").strip()
            if not singlish_input:
                continue

            expected_output = str(top_left_cell(ws, row, expected_col).value or "").strip()
            tested_count += 1
            print(f"[{tested_count}] Row {row}: {singlish_input[:70]}...")

            actual_output = ""
            try:
                for attempt in range(1, args.retries + 1):
                    actual_output = run_single_case(
                        context=context,
                        url=args.url,
                        singlish_input=singlish_input,
                        wait_ms=args.wait_ms,
                        type_delay_ms=args.type_delay_ms,
                        timeout_ms=args.timeout_ms,
                    )

                    if actual_output:
                        break

                    if attempt < args.retries:
                        print(f"    Retry {attempt}: Actual output empty")

                status = "PASS" if actual_output == expected_output else "FAIL"
                top_left_cell(ws, row, actual_col).value = actual_output
                top_left_cell(ws, row, status_col).value = status

                print(f"    Actual: {actual_output[:120] if actual_output else '[EMPTY]'}")
                print(f"    Status: {status}")

                if args.save_every > 0 and tested_count % args.save_every == 0:
                    wb.save(excel_path)

            except Exception as e:
                top_left_cell(ws, row, actual_col).value = f"ERROR: {e}"
                top_left_cell(ws, row, status_col).value = "ERROR"
                print(f"    ERROR: {e}")

        wb.save(excel_path)
        context.close()
        browser.close()

    print(f"\nDone. Results saved to: {excel_path}")


def parse_args():
    parser = argparse.ArgumentParser(description="Run Assignment 1 Singlish transliteration test cases.")
    parser.add_argument("--excel", required=True, help="Excel file path, e.g. Assignment 1 - Test cases.xlsx")
    parser.add_argument("--url", default=DEFAULT_URL)
    parser.add_argument("--sheet", default=None)
    parser.add_argument("--wait-ms", type=int, default=12000)
    parser.add_argument("--type-delay-ms", type=int, default=80)
    parser.add_argument("--slow-mo-ms", type=int, default=150)
    parser.add_argument("--timeout-ms", type=int, default=60000)
    parser.add_argument("--save-every", type=int, default=1)
    parser.add_argument("--retries", type=int, default=2)
    parser.add_argument("--headless", action="store_true", default=False)
    return parser.parse_args()


if __name__ == "__main__":
    configure_stdout()
    run_tests(parse_args())
